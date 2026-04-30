# compras/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from io import BytesIO
from decimal import Decimal
import logging
from django.contrib.auth.decorators import login_required

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook

from .models import Proveedor, Producto, OrdenCompra, Factura
from .forms import ProveedorForm, ProductoForm, OrdenCompraForm, FacturaForm

logger = logging.getLogger(__name__)

# Intentamos import del servicio contable; si no existe, lo ignoramos (no romperá la app)
try:
    from contabilidad.services import generar_asiento_desde_factura
    CONTABILIDAD_AVAILABLE = True
except Exception as e:
    CONTABILIDAD_AVAILABLE = False
    logger.warning("Servicio de contabilidad no disponible: los asientos no se generarán automáticamente. (%s)", str(e))


def dashboard_compras(request):
    context = {
        'titulo_modulo': 'Módulo de Compras',
        'descripcion_modulo': 'Gestión de proveedores, productos, órdenes de compra y facturación.',
        'total_proveedores': Proveedor.objects.count(),
        'total_productos': Producto.objects.count(),
        'total_ordenes': OrdenCompra.objects.count(),
        'total_facturas': Factura.objects.count(),
    }
    return render(request, 'compras/dashboard_compras.html', context)


# ----------------- proveedores, productos, ordenes -----------------
def lista_proveedores(request):
    q = request.GET.get('q', '')
    proveedores = Proveedor.objects.filter(nombre_empresa__icontains=q) if q else Proveedor.objects.all()
    return render(request, 'compras/proveedores/list.html', {'proveedores': proveedores})


def nuevo_proveedor(request):
    form = ProveedorForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Proveedor registrado exitosamente.")
        return redirect('compras:lista_proveedores')
    return render(request, 'compras/proveedores/form.html', {'form': form, 'accion': 'Nuevo Proveedor'})


def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    form = ProveedorForm(request.POST or None, instance=proveedor)
    if form.is_valid():
        form.save()
        messages.success(request, "Proveedor actualizado correctamente.")
        return redirect('compras:lista_proveedores')
    return render(request, 'compras/proveedores/form.html', {'form': form, 'accion': 'Editar Proveedor'})


def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.delete()
        messages.success(request, "Proveedor eliminado correctamente.")
        return redirect('compras:lista_proveedores')
    return render(request, 'compras/proveedores/confirm_delete.html', {'objeto': proveedor})


def lista_productos(request):
    productos = Producto.objects.all()
    return render(request, 'compras/productos/list.html', {'productos': productos})


def nuevo_producto(request):
    form = ProductoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Producto agregado correctamente.")
        return redirect('compras:lista_productos')
    return render(request, 'compras/productos/form.html', {'form': form, 'accion': 'Nuevo Producto'})


def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    if form.is_valid():
        form.save()
        messages.success(request, "Producto actualizado correctamente.")
        return redirect('compras:lista_productos')
    return render(request, 'compras/productos/form.html', {'form': form, 'accion': 'Editar Producto'})


def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.delete()
        messages.success(request, "Producto eliminado correctamente.")
        return redirect('compras:lista_productos')
    return render(request, 'compras/productos/confirm_delete.html', {'objeto': producto})


def lista_ordenes(request):
    ordenes = OrdenCompra.objects.all()
    return render(request, 'compras/ordenes/list.html', {'ordenes': ordenes})


def nueva_orden(request):
    form = OrdenCompraForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Orden de compra registrada correctamente.")
        return redirect('compras:lista_ordenes')
    return render(request, 'compras/ordenes/form.html', {'form': form, 'accion': 'Nueva Orden de Compra'})


def editar_orden(request, pk):
    orden = get_object_or_404(OrdenCompra, pk=pk)
    form = OrdenCompraForm(request.POST or None, instance=orden)
    if form.is_valid():
        form.save()
        messages.success(request, "Orden de compra actualizada correctamente.")
        return redirect('compras:lista_ordenes')
    return render(request, 'compras/ordenes/form.html', {'form': form, 'accion': 'Editar Orden de Compra'})


def eliminar_orden(request, pk):
    orden = get_object_or_404(OrdenCompra, pk=pk)
    if request.method == 'POST':
        orden.delete()
        messages.success(request, "Orden eliminada correctamente.")
        return redirect('compras:lista_ordenes')
    return render(request, 'compras/ordenes/confirm_delete.html', {'objeto': orden})


def ver_orden(request, pk):
    orden = get_object_or_404(OrdenCompra, pk=pk)
    detalles = orden.detalles.all()
    subtotal = sum(d.subtotal for d in detalles)
    iva = subtotal * Decimal('0.16')
    total = subtotal + iva
    context = {'orden': orden, 'detalles': detalles, 'subtotal': subtotal, 'iva': iva, 'total': total}
    return render(request, 'compras/ordenes/ver_orden.html', context)


def imprimir_orden_pdf(request, pk):
    orden = get_object_or_404(OrdenCompra, pk=pk)
    detalles = orden.detalles.all()
    subtotal = sum(d.subtotal for d in detalles)
    iva = subtotal * Decimal('0.16')
    total = subtotal + iva

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.setFont("Helvetica", 11)

    p.drawString(50, 750, f"Orden de Compra N° {orden.numero_oc}")
    p.drawString(50, 735, f"Proveedor: {orden.proveedor.nombre_empresa}")
    p.drawString(50, 720, f"Fecha: {orden.fecha.strftime('%d/%m/%Y')}")

    y = 690
    p.drawString(50, y, "Producto")
    p.drawString(250, y, "Cantidad")
    p.drawString(350, y, "Precio Unitario")
    p.drawString(470, y, "Subtotal")
    y -= 15

    for d in detalles:
        nombre = getattr(d.producto, 'descripcion', str(d.producto))[:35]
        p.drawString(50, y, nombre)
        p.drawString(270, y, str(d.cantidad))
        p.drawRightString(430, y, f"{d.precio_unitario:.2f}")
        p.drawRightString(520, y, f"{d.subtotal:.2f}")
        y -= 15

    y -= 10
    p.line(50, y, 520, y)
    y -= 20
    p.drawRightString(450, y, "Subtotal:")
    p.drawRightString(520, y, f"{subtotal:.2f}")
    y -= 15
    p.drawRightString(450, y, "IVA (16%):")
    p.drawRightString(520, y, f"{iva:.2f}")
    y -= 15
    p.drawRightString(450, y, "Total:")
    p.drawRightString(520, y, f"{total:.2f}")

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="orden_{orden.numero_oc}.pdf"'
    return response


def exportar_orden_excel(request, pk):
    orden = get_object_or_404(OrdenCompra, pk=pk)
    detalles = orden.detalles.all()
    subtotal = sum(d.subtotal for d in detalles)
    iva = subtotal * Decimal('0.16')
    total = subtotal + iva

    wb = Workbook()
    ws = wb.active
    ws.title = f"OC_{orden.numero_oc}"

    ws['A1'] = f"Orden de Compra N° {orden.numero_oc}"
    ws['A2'] = f"Proveedor: {orden.proveedor.nombre_empresa}"
    ws['A3'] = f"Fecha: {orden.fecha.strftime('%d/%m/%Y')}"

    ws.append([])
    ws.append(["Producto", "Cantidad", "Precio Unitario (Bs.)", "Subtotal (Bs.)"])

    for d in detalles:
        ws.append([d.producto.descripcion, d.cantidad, float(d.precio_unitario), float(d.subtotal)])

    ws.append([])
    ws.append(["", "", "Subtotal", float(subtotal)])
    ws.append(["", "", "IVA (16%)", float(iva)])
    ws.append(["", "", "Total", float(total)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=orden_{orden.numero_oc}.xlsx'
    wb.save(response)
    return response


# ----------------- FACTURAS -----------------
def buscar_proveedor(request):
    rif = request.GET.get('rif')
    if not rif:
        return JsonResponse({'error': 'RIF requerido'}, status=400)
    try:
        proveedor = Proveedor.objects.get(rif__iexact=rif.strip())
        return JsonResponse({'id': proveedor.id, 'nombre_empresa': proveedor.nombre_empresa, 'rif': proveedor.rif})
    except Proveedor.DoesNotExist:
        return JsonResponse({'error': 'Proveedor no encontrado'}, status=404)


def lista_facturas(request):
    facturas = Factura.objects.select_related('proveedor').all()
    return render(request, 'compras/facturas/list.html', {'facturas': facturas})


def crear_factura(request):
    if request.method == 'POST':
        form = FacturaForm(request.POST)
        if form.is_valid():
            factura = form.save(commit=False)

            # Resolver proveedor por rif_proveedor si fue digitado
            rif = form.cleaned_data.get('rif_proveedor') or ''
            if rif:
                try:
                    proveedor = Proveedor.objects.get(rif__iexact=rif.strip())
                    factura.proveedor = proveedor
                except Proveedor.DoesNotExist:
                    messages.error(request, f"No se encontró un proveedor con el RIF {rif}.")
                    return render(request, 'compras/facturas/form.html', {'form': form, 'accion': 'Crear Factura'})

            # Si no hay proveedor asignado, usar el select
            if not getattr(factura, 'proveedor', None):
                if form.cleaned_data.get('proveedor'):
                    factura.proveedor = form.cleaned_data.get('proveedor')
                else:
                    messages.error(request, "Debe seleccionar o buscar un proveedor.")
                    return render(request, 'compras/facturas/form.html', {'form': form, 'accion': 'Crear Factura'})

            # Recalcular montos en backend (autoridad)
            monto_base = form.cleaned_data.get('monto_base') or Decimal('0.00')
            tasa = form.cleaned_data.get('tasa_cambio') or Decimal('1.00')
            moneda = form.cleaned_data.get('moneda')
            aplica_iva = form.cleaned_data.get('aplica_iva') if 'aplica_iva' in form.cleaned_data else True

            base_en_bs = (monto_base * tasa) if moneda == 'USD' else monto_base

            iva = (base_en_bs * Decimal('0.16')) if aplica_iva else Decimal('0.00')
            total = (base_en_bs + iva).quantize(Decimal('0.01'))

            # retenciones porcentuales (los pct vienen del form extras)
            ret_iva_pct = (form.cleaned_data.get('ret_iva_pct') or Decimal('0.00')) / Decimal('100.00')
            ret_islr_pct = (form.cleaned_data.get('ret_islr_pct') or Decimal('0.00')) / Decimal('100.00')
            ret_mun_pct = (form.cleaned_data.get('ret_municipal_pct') or Decimal('0.00')) / Decimal('100.00')

            ret_iva = (base_en_bs * ret_iva_pct).quantize(Decimal('0.01'))
            ret_islr = (base_en_bs * ret_islr_pct).quantize(Decimal('0.01'))
            ret_mun = (base_en_bs * ret_mun_pct).quantize(Decimal('0.01'))

            factura.iva = iva.quantize(Decimal('0.01'))
            factura.total_factura = total
            factura.ret_iva = ret_iva
            factura.ret_islr = ret_islr
            factura.porcentaje_ret_municipal = (form.cleaned_data.get('ret_municipal_pct') or Decimal('0.00'))
            factura.ret_municipal = ret_mun

            factura.save()

            # Generar asiento en contabilidad (si disponible)
            if CONTABILIDAD_AVAILABLE:
                try:
                    generar_asiento_desde_factura(factura)
                except Exception as e:
                    # No rompemos la transacción principal; registramos error para revisión
                    logger.exception("Error generando asiento para factura %s: %s", factura.pk, str(e))
                    messages.warning(request, "Factura guardada, pero no se pudo generar el asiento contable automáticamente (revise logs).")

            messages.success(request, "Factura registrada correctamente.")
            return redirect('compras:lista_facturas')
        else:
            # form inválido
            messages.error(request, "Corrige los errores del formulario.")
            return render(request, 'compras/facturas/form.html', {'form': form, 'accion': 'Crear Factura'})
    else:
        form = FacturaForm()
    return render(request, 'compras/facturas/form.html', {'form': form, 'accion': 'Crear Factura'})


def editar_factura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    if request.method == 'POST':
        form = FacturaForm(request.POST, instance=factura)
        if form.is_valid():
            factura = form.save(commit=False)
            # Recalculos idem a crear
            monto_base = form.cleaned_data.get('monto_base') or Decimal('0.00')
            tasa = form.cleaned_data.get('tasa_cambio') or Decimal('1.00')
            moneda = form.cleaned_data.get('moneda')
            aplica_iva = form.cleaned_data.get('aplica_iva') if 'aplica_iva' in form.cleaned_data else True

            base_en_bs = (monto_base * tasa) if moneda == 'USD' else monto_base
            iva = (base_en_bs * Decimal('0.16')) if aplica_iva else Decimal('0.00')
            total = (base_en_bs + iva).quantize(Decimal('0.01'))

            ret_iva_pct = (form.cleaned_data.get('ret_iva_pct') or Decimal('0.00')) / Decimal('100.00')
            ret_islr_pct = (form.cleaned_data.get('ret_islr_pct') or Decimal('0.00')) / Decimal('100.00')
            ret_mun_pct = (form.cleaned_data.get('ret_municipal_pct') or Decimal('0.00')) / Decimal('100.00')

            ret_iva = (base_en_bs * ret_iva_pct).quantize(Decimal('0.01'))
            ret_islr = (base_en_bs * ret_islr_pct).quantize(Decimal('0.01'))
            ret_mun = (base_en_bs * ret_mun_pct).quantize(Decimal('0.01'))

            factura.iva = iva.quantize(Decimal('0.01'))
            factura.total_factura = total
            factura.ret_iva = ret_iva
            factura.ret_islr = ret_islr
            factura.porcentaje_ret_municipal = (form.cleaned_data.get('ret_municipal_pct') or Decimal('0.00'))
            factura.ret_municipal = ret_mun

            factura.save()

            # Regenerar asiento: decisión de negocio (aquí, intentamos generar uno nuevo)
            if CONTABILIDAD_AVAILABLE:
                try:
                    generar_asiento_desde_factura(factura)
                except Exception as e:
                    logger.exception("Error regenerando asiento para factura %s: %s", factura.pk, str(e))
                    messages.warning(request, "Factura actualizada, pero no se pudo (re)generar el asiento contable automáticamente (revise logs).")

            messages.success(request, "Factura actualizada correctamente.")
            return redirect('compras:lista_facturas')
        else:
            messages.error(request, "Corrige los errores del formulario.")
            return render(request, 'compras/facturas/form.html', {'form': form, 'accion': 'Editar Factura'})
    else:
        form = FacturaForm(instance=factura)
        # inicializamos rif_proveedor para que se muestre en el campo de texto
        form.fields['rif_proveedor'].initial = factura.proveedor.rif if factura.proveedor else ''
        # inicializamos los pct de retención para que aparezcan en inputs no-model
        form.fields['ret_iva_pct'].initial = Decimal('0.00')
        form.fields['ret_islr_pct'].initial = Decimal('0.00')
        form.fields['ret_municipal_pct'].initial = factura.porcentaje_ret_municipal or Decimal('0.00')

    return render(request, 'compras/facturas/form.html', {'form': form, 'accion': 'Editar Factura'})


def eliminar_factura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    if request.method == 'POST':
        factura.delete()
        messages.success(request, "Factura eliminada correctamente.")
        return redirect('compras:lista_facturas')
    return render(request, 'compras/facturas/confirm_delete.html', {'objeto': factura})


def facturas_pendientes(request):
    return render(request, "compras/facturas_pendientes.html")
