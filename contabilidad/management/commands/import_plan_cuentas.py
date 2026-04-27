# contabilidad/management/commands/import_plan_cuentas.py
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from contabilidad.models import CuentaContable

class Command(BaseCommand):
    help = 'Importa plan de cuentas desde un archivo Excel. Uso: manage.py import_plan_cuentas --file /mnt/data/Plan de Cuentas.xls'

    def add_arguments(self, parser):
        parser.add_argument('file_path', nargs='?', type=str, help='Ruta al archivo Excel (por defecto /mnt/data/Plan de Cuentas.xls)')

    def handle(self, *args, **options):
        file_path = options.get('file_path') or '/mnt/data/Plan de Cuentas.xls'
        try:
            wb = load_workbook(filename=file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"No se pudo abrir el archivo: {e}")

        sheet = wb.active

        # Asumimos que la primera fila contiene encabezados y que las columnas incluyen:
        # Col A: Codigo, Col B: Cuenta Contable, Col C.. son flags opcionales
        created = 0
        updated = 0
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            codigo = (row[0].value or '').strip() if row[0].value else ''
            nombre = (row[1].value or '').strip() if row[1].value else ''
            if not codigo or not nombre:
                continue

            # Deduzco nivel por cantidad de puntos en el código o por estructura
            nivel = codigo.count('.') + 1 if '.' in codigo else len(codigo.split('-')) if '-' in codigo else 1
            naturaleza = 'D'  # default; podrías leer columna para esto
            cuenta, created_flag = CuentaContable.objects.update_or_create(
                codigo=codigo,
                defaults={'nombre': nombre, 'nivel': nivel, 'naturaleza': naturaleza, 'es_activo': True}
            )
            if created_flag:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(f'Importación finalizada. Creadas: {created}. Actualizadas: {updated}.'))
